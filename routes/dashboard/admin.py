from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app
from models.baseDatos import db, nuevaHabitacion, Usuario, InventarioHabitacion, InventarioItem, Post
from datetime import datetime
from flask import session
from flask import send_file, make_response
import io, csv
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    def get_column_letter(col_index:int):
        # Fallback simple mapping A..Z, AA.. as naive
        letters = ''
        i = col_index
        while i:
            i, rem = divmod(i - 1, 26)
            letters = chr(65 + rem) + letters
        return letters or 'A'
from datetime import date
try:
    # ReportLab para exportar PDF del inventario
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
except Exception:
    # Si no está instalado en dev, la ruta seguirá fallando hasta instalar requirements
    letter = None

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

# Restringir el acceso a administradores para todas las rutas de este blueprint
@admin_bp.before_request
def _require_admin():
    user = session.get('user')
    if not user or user.get('rol') != 'admin':
        flash('Acceso restringido solo para administradores', 'warning')
        return redirect(url_for('registro.login'))

# Ruta para mostrar formulario de edición de habitación
@admin_bp.route("/hospedaje/editar/<int:habitacion_id>", methods=["GET"])
def hospedaje_editar(habitacion_id):
    habitacion = nuevaHabitacion.query.get_or_404(habitacion_id)
    return render_template("dashboard/editar_habitacion.html", habitacion=habitacion)

# Ruta para actualizar habitación en la base de datos
@admin_bp.route("/hospedaje/editar/<int:habitacion_id>", methods=["POST"])
def hospedaje_actualizar(habitacion_id):
    habitacion = nuevaHabitacion.query.get_or_404(habitacion_id)
    try:
        habitacion.nombre = request.form["nombre"]
        # Asegurar que la descripción también se actualiza
        habitacion.descripcion = request.form.get("descripcion", habitacion.descripcion)
        habitacion.plan = request.form.get("plan") or None
        try:
            habitacion.numero = int(request.form.get("numero")) if request.form.get("numero") else None
        except Exception:
            habitacion.numero = None
        habitacion.caracteristicas = request.form.get("caracteristicas") or habitacion.caracteristicas
        habitacion.precio = float(request.form["precio"])
        habitacion.cupo_personas = int(request.form.get("cupo_personas", 1))
        habitacion.estado = request.form.get("estado", "Disponible")
        imagen_file = request.files.get("imagen")
        if imagen_file and imagen_file.filename:
            import os, time
            from uuid import uuid4
            from werkzeug.utils import secure_filename
            filename = secure_filename(imagen_file.filename)
            unique = f"{int(time.time())}_{uuid4().hex[:8]}_{filename}"
            img_folder = os.path.join(current_app.static_folder, "img", "uploads")
            os.makedirs(img_folder, exist_ok=True)
            save_path = os.path.join(img_folder, unique)
            imagen_file.save(save_path)
            habitacion.imagen = f"img/uploads/{unique}"
        db.session.commit()
        flash("✅ Habitación actualizada correctamente", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al actualizar la habitación: {e}", "danger")
    return redirect(url_for("admin.hospedaje_index"))
# (Eliminar duplicado de admin_bp e imports redundantes) 

# ==========================
# 📌 SECCIÓN HOSPEDAJE
# ==========================
@admin_bp.route("/hospedaje")
def hospedaje_index():
    habitaciones = nuevaHabitacion.query.all()
    return render_template("dashboard/hospedaje_admin.html", habitaciones=habitaciones)

# ==========================
# SECCIÓN INICIO (Home) - CRUD de contenido usando Post
# ==========================
def _save_uploaded_image(file_field_name: str):
    """Guarda una imagen subida en instance/uploads y devuelve la ruta relativa 'uploads/<file>' o None."""
    img = request.files.get(file_field_name)
    if img and getattr(img, 'filename', ''):
        from werkzeug.utils import secure_filename
        import os, time
        from uuid import uuid4
        filename = secure_filename(img.filename)
        unique = f"{int(time.time())}_{uuid4().hex[:8]}_{filename}"
        # Guardar siempre en instance/uploads (persistente y servido por /media)
        inst_uploads = os.path.join(current_app.instance_path, 'uploads')
        os.makedirs(inst_uploads, exist_ok=True)
        save_path = os.path.join(inst_uploads, unique)
        img.save(save_path)
        return f"uploads/{unique}"
    return None

@admin_bp.route('/home/post/create', methods=['POST'])
def home_post_create():
    try:
        titulo = request.form.get('titulo') or 'Sin título'
        contenido = request.form.get('contenido') or ''
        imagen_path = _save_uploaded_image('imagen')
        # Calcular siguiente orden dentro de 'home'
        last = Post.query.filter_by(categoria='home').order_by(Post.orden.desc()).first()
        next_order = (last.orden + 1) if last else 1
        post = Post(titulo=titulo, contenido=contenido, categoria='home', orden=next_order)
        if imagen_path:
            post.imagen = imagen_path
        db.session.add(post)
        db.session.commit()
        flash('Sección creada correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'No se pudo crear la sección: {e}', 'danger')
    return redirect(url_for('main.home_admin'))

@admin_bp.route('/home/post/<int:post_id>/update', methods=['POST'])
def home_post_update(post_id):
    post = Post.query.get_or_404(post_id)
    try:
        post.titulo = request.form.get('titulo') or post.titulo
        post.contenido = request.form.get('contenido') or post.contenido
        new_img = _save_uploaded_image('imagen')
        if new_img:
            post.imagen = new_img
        db.session.commit()
        flash('Sección actualizada', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar: {e}', 'danger')
    return redirect(url_for('main.home_admin'))

@admin_bp.route('/home/post/<int:post_id>/delete', methods=['POST'])
def home_post_delete(post_id):
    post = Post.query.get_or_404(post_id)
    try:
        db.session.delete(post)
        db.session.commit()
        flash('Sección eliminada', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'No se pudo eliminar: {e}', 'danger')
    return redirect(url_for('main.home_admin'))

@admin_bp.route('/home/post/<int:post_id>/orden', methods=['POST'])
def home_post_orden(post_id):
    """Mueve un post arriba o abajo intercambiando su 'orden' con el vecino."""
    direction = request.form.get('dir')  # 'up' or 'down'
    post = Post.query.get_or_404(post_id)
    if post.categoria != 'home':
        flash('Operación inválida', 'danger')
        return redirect(url_for('main.home_admin'))
    try:
        if direction == 'up':
            neighbor = Post.query.filter_by(categoria='home').filter(Post.orden < post.orden).order_by(Post.orden.desc()).first()
        else:
            neighbor = Post.query.filter_by(categoria='home').filter(Post.orden > post.orden).order_by(Post.orden.asc()).first()
        if neighbor:
            post.orden, neighbor.orden = neighbor.orden, post.orden
            db.session.commit()
        else:
            flash('No hay más elementos para reordenar', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'No se pudo reordenar: {e}', 'danger')
    return redirect(url_for('main.home_admin'))

# ==========================
# 📄 INVENTARIO DE HABITACIÓN (vista)
# ==========================
@admin_bp.route("/inventario")
def inventario_view():
    room_id = request.args.get('room_id', type=int)
    rec_id = request.args.get('rec_id', type=int)
    # Permitir activar auto-PDF con ?pdf=1 o ?auto_pdf=1
    auto_pdf = False
    for key in ('pdf', 'auto_pdf'):
        val = request.args.get(key)
        if val and str(val).lower() in ('1','true','yes','on'):
            auto_pdf = True
            break
    habitacion = None
    record = None
    items_map = {}
    if room_id:
        habitacion = nuevaHabitacion.query.get(room_id)
    if rec_id:
        record = InventarioHabitacion.query.get_or_404(rec_id)
        # construir mapa de items por clave
        for it in record.items:
            items_map[it.key] = {
                'checked': bool(it.checked),
                'quantity': it.quantity,
                'value_text': it.value_text,
            }
    hotel_name = "Hotel Isla Encanto"
    return render_template(
        "dashboard/inventario.html",
        habitacion=habitacion,
        hotel_name=hotel_name,
        record=record,
        items_map=items_map,
        auto_pdf=auto_pdf,
    )

@admin_bp.route("/inventario", methods=["POST"])
def inventario_save():
    form = request.form
    rec_id = form.get('rec_id', type=int)
    # Helper to parse date 'YYYY-MM-DD' -> date
    def _pdate(val):
        try:
            return datetime.strptime(val, '%Y-%m-%d').date() if val else None
        except Exception:
            return None

    if rec_id:
        rec = InventarioHabitacion.query.get_or_404(rec_id)
        # actualizar campos del registro existente
        rec.habitacion_id = (int(form.get('habitacion_id')) if form.get('habitacion_id') else None)
        rec.hotel = form.get('hotel')
        rec.room_number = form.get('room_number')
        rec.room_type = form.get('room_type')
        rec.inspection_date = _pdate(form.get('inspection_date'))
        rec.inspector = form.get('inspector')
        rec.observations = form.get('observations')
        rec.rating_cleaning = int(form.get('rating_cleaning', 0) or 0)
        rec.rating_furniture = int(form.get('rating_furniture', 0) or 0)
        rec.rating_equipment = int(form.get('rating_equipment', 0) or 0)
        rec.inspector_signature = form.get('inspector_signature')
        rec.inspector_date = _pdate(form.get('inspector_date'))
        rec.supervisor_signature = form.get('supervisor_signature')
        rec.supervisor_date = _pdate(form.get('supervisor_date'))
        # limpiar items previos
        for it in list(rec.items):
            db.session.delete(it)
    else:
        rec = InventarioHabitacion(
            habitacion_id = (int(form.get('habitacion_id')) if form.get('habitacion_id') else None),
            hotel = form.get('hotel'),
            room_number = form.get('room_number'),
            room_type = form.get('room_type'),
            inspection_date = _pdate(form.get('inspection_date')),
            inspector = form.get('inspector'),
            observations = form.get('observations'),
            rating_cleaning = int(form.get('rating_cleaning', 0) or 0),
            rating_furniture = int(form.get('rating_furniture', 0) or 0),
            rating_equipment = int(form.get('rating_equipment', 0) or 0),
            inspector_signature = form.get('inspector_signature'),
            inspector_date = _pdate(form.get('inspector_date')),
            supervisor_signature = form.get('supervisor_signature'),
            supervisor_date = _pdate(form.get('supervisor_date')),
        )

    # Parse dynamic items: expect names like item__category__key fields: check, qty, text
    # For simplicity, define a fixed list mapping to our form elements
    def add_item(category, key, label, checked, quantity=None, value_text=None):
        itm = InventarioItem(
            record=rec,
            category=category,
            key=key,
            label=label,
            checked=bool(checked),
            quantity=(int(quantity) if (quantity is not None and str(quantity).strip() != '') else None),
            value_text=(value_text if value_text else None)
        )
        db.session.add(itm)

    # Map of fields: (category, checkbox name, qty name, text name, label)
    mappings = [
        ("dorm_mob", 'bed', 'bed_qty', None, 'Cama'),
        ("dorm_mob", 'nightstand', 'nightstand_qty', None, 'Mesitas de noche'),
        ("dorm_mob", 'lamps', 'lamps_qty', None, 'Lámparas de mesa'),
        ("dorm_mob", 'closet', None, None, 'Armario/Closet'),
        ("dorm_mob", 'chair', 'chair_qty', None, 'Silla/Sillón'),
        ("dorm_mob", 'desk', None, None, 'Escritorio'),
        ("dorm_mob", 'mirror', None, None, 'Espejo de pared'),
        ("ropa", 'mattress', None, 'mattress_state', 'Colchón'),
        ("ropa", 'pillows', 'pillows_qty', None, 'Almohadas'),
        ("ropa", 'sheets', None, None, 'Sábanas juego completo'),
        ("ropa", 'comforter', None, None, 'Edredón/Colcha'),
        ("ropa", 'blanket', None, None, 'Cobertor adicional'),
        ("ropa", 'cushions', 'cushions_qty', None, 'Cojines decorativos'),
        ("banio", 'bath_mirror', None, None, 'Espejo'),
        ("banio", 'towel_rack', None, None, 'Toallero/Perchero'),
        ("banio", 'trash_bin', None, None, 'Papelera'),
        ("banio", 'scale', None, None, 'Balanza/Báscula'),
        ("banio", 'hairdryer', None, None, 'Secador de cabello'),
        ("banio_am", 'bath_towels', 'bath_towels_qty', None, 'Toallas grandes'),
        ("banio_am", 'hand_towels', 'hand_towels_qty', None, 'Toallas de mano'),
        ("banio_am", 'soap', None, None, 'Jabón/Gel de baño'),
        ("banio_am", 'shampoo', None, None, 'Shampoo'),
        ("banio_am", 'conditioner', None, None, 'Acondicionador'),
        ("banio_am", 'lotion', None, None, 'Loción corporal'),
        ("banio_am", 'dental_kit', None, None, 'Kit dental'),
        ("banio_am", 'toilet_paper', 'toilet_paper_qty', None, 'Papel higiénico'),
        ("electro", 'tv', None, 'tv_model', 'Televisor'),
        ("electro", 'remote', None, None, 'Control remoto TV'),
        ("electro", 'phone', None, None, 'Teléfono'),
        ("electro", 'alarm', None, None, 'Radio reloj despertador'),
        ("electro", 'safe', None, None, 'Caja fuerte'),
        ("electro", 'minibar', None, None, 'Minibar/Refrigerador'),
        ("electro", 'coffee_maker', None, None, 'Cafetera/Hervidor'),
        ("menaje", 'glasses', 'glasses_qty', None, 'Vasos de vidrio'),
        ("menaje", 'cups', 'cups_qty', None, 'Tazas/Mugs'),
        ("menaje", 'spoons', 'spoons_qty', None, 'Cucharas'),
        ("menaje", 'opener', None, None, 'Abrebotellas'),
        ("menaje", 'coffee', 'coffee_qty', None, 'Café/Té'),
        ("menaje", 'sugar', 'sugar_qty', None, 'Azúcar/Endulzante'),
        ("otros", 'curtains', None, None, 'Cortinas decorativas'),
        ("otros", 'blackout', None, None, 'Cortinas blackout'),
        ("otros", 'directory', None, None, 'Directorio de servicios'),
        ("otros", 'menu', None, None, 'Menú room service'),
        ("otros", 'smoke_detector', None, None, 'Detector de humo'),
        ("otros", 'evacuation_map', None, None, 'Mapa de evacuación'),
        ("otros", 'hangers', 'hangers_qty', None, 'Perchas en closet'),
        ("otros", 'umbrella', None, None, 'Paraguas'),
    ]

    try:
        if not rec_id:
            db.session.add(rec)
        for cat, key, qty_key, text_key, label in mappings:
            checked = form.get(key) == 'on'
            qty = form.get(qty_key) if qty_key else None
            text_val = form.get(text_key) if text_key else None
            add_item(cat, key, label, checked, qty, text_val)
        db.session.commit()
        flash("✅ Inventario guardado" if not rec_id else "✅ Inventario actualizado", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error guardando inventario: {e}", "danger")
    return redirect(url_for('admin.inventarios_list'))

@admin_bp.route('/inventarios')
def inventarios_list():
    # Filtros por parámetros GET
    room = (request.args.get('room') or '').strip()
    rtype = (request.args.get('type') or '').strip()
    inspector = (request.args.get('inspector') or '').strip()
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    q = InventarioHabitacion.query
    if room:
        q = q.filter(InventarioHabitacion.room_number.ilike(f"%{room}%"))
    if rtype:
        q = q.filter(InventarioHabitacion.room_type.ilike(f"%{rtype}%"))
    if inspector:
        q = q.filter(InventarioHabitacion.inspector.ilike(f"%{inspector}%"))
    # Rango de fechas de inspección
    def _pdate(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date() if s else None
        except Exception:
            return None
    d_from = _pdate(date_from)
    d_to = _pdate(date_to)
    if d_from:
        q = q.filter(InventarioHabitacion.inspection_date >= d_from)
    if d_to:
        q = q.filter(InventarioHabitacion.inspection_date <= d_to)

    q = q.order_by(InventarioHabitacion.created_at.desc())
    registros = q.limit(200).all()
    total_count = q.count()
    activos_count = q.filter(InventarioHabitacion.inspection_date.isnot(None)).count()

    # Listas distintas para selects
    # Nota: usamos todo el conjunto (sin aplicar filtros) para opciones completas
    all_types = [row[0] for row in db.session.query(InventarioHabitacion.room_type).filter(InventarioHabitacion.room_type.isnot(None)).distinct().order_by(InventarioHabitacion.room_type.asc()).all()]
    all_inspectors = [row[0] for row in db.session.query(InventarioHabitacion.inspector).filter(InventarioHabitacion.inspector.isnot(None)).distinct().order_by(InventarioHabitacion.inspector.asc()).all()]

    return render_template(
        'dashboard/inventarios_list.html',
        registros=registros,
        total_count=total_count,
        activos_count=activos_count,
        room=room,
        rtype=rtype,
        inspector=inspector,
        date_from=date_from or '',
        date_to=date_to or '',
        types_list=all_types,
        inspectors_list=all_inspectors,
    )

@admin_bp.route('/inventario/export/csv')
def inventario_export_csv():
    """Exporta un registro de inventario a CSV (compatible con Excel)."""
    rec_id = request.args.get('rec_id', type=int)
    if not rec_id:
        flash('Falta rec_id', 'warning')
        return redirect(url_for('admin.inventarios_list'))
    rec = InventarioHabitacion.query.get_or_404(rec_id)

    output = io.StringIO()
    writer = csv.writer(output)

    # Encabezado de registro
    writer.writerow(['Inventario ID', rec.id])
    writer.writerow(['Hotel', rec.hotel or ''])
    writer.writerow(['Habitación', rec.room_number or ''])
    writer.writerow(['Tipo', rec.room_type or ''])
    writer.writerow(['Fecha inspección', rec.inspection_date or ''])
    writer.writerow(['Inspector', rec.inspector or ''])
    writer.writerow(['Observaciones', (rec.observations or '').replace('\n', ' ')])
    writer.writerow(['Limpieza (1-5)', rec.rating_cleaning or 0])
    writer.writerow(['Mobiliario (1-5)', rec.rating_furniture or 0])
    writer.writerow(['Equipos (1-5)', rec.rating_equipment or 0])
    writer.writerow(['Firma Inspector', rec.inspector_signature or ''])
    writer.writerow(['Fecha Inspector', rec.inspector_date or ''])
    writer.writerow(['Firma Supervisor', rec.supervisor_signature or ''])
    writer.writerow(['Fecha Supervisor', rec.supervisor_date or ''])
    writer.writerow([])

    # Items
    writer.writerow(['Categoria', 'Clave', 'Etiqueta', 'Marcado', 'Cantidad', 'Texto'])
    for it in (rec.items or []):
        writer.writerow([
            it.category or '',
            it.key or '',
            it.label or '',
            'Sí' if it.checked else 'No',
            it.quantity if it.quantity is not None else '',
            (it.value_text or ''),
        ])

    csv_bytes = output.getvalue().encode('utf-8-sig')  # BOM para Excel
    mem = io.BytesIO(csv_bytes)
    filename = f"Inventario_{(rec.room_number or 'Habitacion')}.csv"
    return send_file(mem, mimetype='text/csv', as_attachment=True, download_name=filename)

@admin_bp.route('/inventario/export/xlsx')
def inventario_export_xlsx():
    rec_id = request.args.get('rec_id', type=int)
    if not rec_id:
        flash('Falta rec_id', 'warning')
        return redirect(url_for('admin.inventarios_list'))
    rec = InventarioHabitacion.query.get_or_404(rec_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    # Encabezado general
    rows = [
        ['Inventario ID', rec.id],
        ['Hotel', rec.hotel or ''],
        ['Habitación', rec.room_number or ''],
        ['Tipo', rec.room_type or ''],
        ['Fecha inspección', str(rec.inspection_date or '')],
        ['Inspector', rec.inspector or ''],
        ['Observaciones', rec.observations or ''],
        ['Limpieza (1-5)', rec.rating_cleaning or 0],
        ['Mobiliario (1-5)', rec.rating_furniture or 0],
        ['Equipos (1-5)', rec.rating_equipment or 0],
        ['Firma Inspector', rec.inspector_signature or ''],
        ['Fecha Inspector', str(rec.inspector_date or '')],
        ['Firma Supervisor', rec.supervisor_signature or ''],
        ['Fecha Supervisor', str(rec.supervisor_date or '')],
        []
    ]
    for r in rows:
        ws.append(r)

    ws.append(['Categoria', 'Clave', 'Etiqueta', 'Marcado', 'Cantidad', 'Texto'])
    for it in (rec.items or []):
        ws.append([
            it.category or '',
            it.key or '',
            it.label or '',
            'Sí' if it.checked else 'No',
            it.quantity if it.quantity is not None else '',
            it.value_text or ''
        ])

    # Auto ancho de columnas básico
    for col in ws.columns:
        max_len = 10
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(60, max_len + 2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"Inventario_{(rec.room_number or 'Habitacion')}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@admin_bp.route('/inventario/export/pdf')
def inventario_export_pdf():
    """Genera un PDF del inventario (servidor) similar al Excel."""
    rec_id = request.args.get('rec_id', type=int)
    if not rec_id:
        flash('Falta rec_id', 'warning')
        return redirect(url_for('admin.inventarios_list'))
    rec = InventarioHabitacion.query.get_or_404(rec_id)

    if letter is None:
        flash('Exportación a PDF no disponible. Instala dependencias.', 'danger')
        return redirect(url_for('admin.inventarios_list'))

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    left = 50
    top = height - 50
    line_height = 14

    def writeln(text, bold=False):
        nonlocal top
        if top < 60:
            c.showPage()
            top = height - 50
        if bold:
            c.setFont('Helvetica-Bold', 11)
        else:
            c.setFont('Helvetica', 10)
        c.drawString(left, top, str(text))
        top -= line_height

    # Encabezado
    c.setTitle(f"Inventario {rec.room_number or ''}")
    c.setFont('Helvetica-Bold', 14)
    c.drawString(left, top, 'Inventario de Habitación')
    top -= 24

    writeln(f"Hotel: {rec.hotel or ''}")
    writeln(f"Habitación: {rec.room_number or ''}")
    writeln(f"Tipo: {rec.room_type or ''}")
    writeln(f"Fecha inspección: {rec.inspection_date or ''}")
    writeln(f"Inspector: {rec.inspector or ''}")
    writeln(f"Observaciones: {(rec.observations or '').replace('\n',' ')}")
    writeln(f"Limpieza (1-5): {rec.rating_cleaning or 0}")
    writeln(f"Mobiliario (1-5): {rec.rating_furniture or 0}")
    writeln(f"Equipos (1-5): {rec.rating_equipment or 0}")
    writeln("", False)
    writeln("Items:", True)
    writeln("Categoria | Clave | Etiqueta | Marcado | Cantidad | Texto", True)

    for it in (rec.items or []):
        row = f"{it.category or ''} | {it.key or ''} | {it.label or ''} | {'Sí' if it.checked else 'No'} | {it.quantity if it.quantity is not None else ''} | {it.value_text or ''}"
        # dividir si es muy largo
        max_chars = 110
        if len(row) <= max_chars:
            writeln(row)
        else:
            # wrap manual simple
            part = row
            while len(part) > 0:
                writeln(part[:max_chars])
                part = part[max_chars:]

    c.showPage()
    c.save()
    buffer.seek(0)
    filename = f"Inventario_{(rec.room_number or 'Habitacion')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

@admin_bp.route('/inventarios/export/xlsx', methods=['POST'])
def inventarios_export_xlsx():
    # Obtener múltiples IDs desde el formulario (name=ids)
    ids = request.form.getlist('ids')
    if not ids:
        flash('Selecciona al menos un inventario', 'warning')
        return redirect(url_for('admin.inventarios_list'))

    # Convertir a int y cargar registros
    try:
        int_ids = [int(x) for x in ids]
    except Exception:
        flash('IDs inválidos', 'danger')
        return redirect(url_for('admin.inventarios_list'))

    recs = InventarioHabitacion.query.filter(InventarioHabitacion.id.in_(int_ids)).order_by(InventarioHabitacion.id.asc()).all()
    if not recs:
        flash('No se encontraron registros', 'warning')
        return redirect(url_for('admin.inventarios_list'))

    wb = Workbook()
    # Eliminar hoja por defecto si vamos a crear varias
    ws0 = wb.active
    ws0.title = 'Resumen'
    ws0.append(['ID', 'Hotel', 'Habitación', 'Tipo', 'Fecha', 'Inspector', 'Creado'])
    for rec in recs:
        ws0.append([
            rec.id,
            rec.hotel or '',
            rec.room_number or '',
            rec.room_type or '',
            str(rec.inspection_date or ''),
            rec.inspector or '',
            str(rec.created_at)
        ])

    def fill_sheet(ws, rec):
        rows = [
            ['Inventario ID', rec.id],
            ['Hotel', rec.hotel or ''],
            ['Habitación', rec.room_number or ''],
            ['Tipo', rec.room_type or ''],
            ['Fecha inspección', str(rec.inspection_date or '')],
            ['Inspector', rec.inspector or ''],
            ['Observaciones', rec.observations or ''],
            ['Limpieza (1-5)', rec.rating_cleaning or 0],
            ['Mobiliario (1-5)', rec.rating_furniture or 0],
            ['Equipos (1-5)', rec.rating_equipment or 0],
            ['Firma Inspector', rec.inspector_signature or ''],
            ['Fecha Inspector', str(rec.inspector_date or '')],
            ['Firma Supervisor', rec.supervisor_signature or ''],
            ['Fecha Supervisor', str(rec.supervisor_date or '')],
            []
        ]
        for r in rows: ws.append(r)
        ws.append(['Categoria', 'Clave', 'Etiqueta', 'Marcado', 'Cantidad', 'Texto'])
        for it in (rec.items or []):
            ws.append([
                it.category or '', it.key or '', it.label or '', 'Sí' if it.checked else 'No', it.quantity if it.quantity is not None else '', it.value_text or ''
            ])
        for col in ws.columns:
            max_len = 10
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try: max_len = max(max_len, len(str(cell.value)))
                except Exception: pass
            ws.column_dimensions[col_letter].width = min(60, max_len + 2)

    # Crear una hoja por registro
    for rec in recs:
        title = f"Inv_{rec.id}"
        ws = wb.create_sheet(title[:31])
        fill_sheet(ws, rec)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"Inventarios_{len(recs)}_registros.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#añadir nueva habitacion ---------------------------------------------------------

@admin_bp.route("/hospedaje/nueva", methods=["POST"])
def hospedaje_nueva():
    try:
        nombre = request.form["nombre"]
        descripcion = request.form["descripcion"]
        print(f"[DEBUG] Descripción recibida: {descripcion}")
        precio = float(request.form["precio"])
        cupo_personas = int(request.form.get("cupo_personas", 1))
        estado = request.form.get("estado", "Disponible")
        plan = request.form.get("plan") or None
        try:
            numero = int(request.form.get("numero")) if request.form.get("numero") else None
        except Exception:
            numero = None
        caracteristicas = request.form.get("caracteristicas") or None
        imagen_file = request.files.get("imagen")
        imagen_path = None
        if imagen_file and imagen_file.filename:
            import os, time
            from uuid import uuid4
            from werkzeug.utils import secure_filename
            filename = secure_filename(imagen_file.filename)
            unique = f"{int(time.time())}_{uuid4().hex[:8]}_{filename}"
            img_folder = os.path.join(current_app.static_folder, "img", "uploads")
            os.makedirs(img_folder, exist_ok=True)
            save_path = os.path.join(img_folder, unique)
            imagen_file.save(save_path)
            imagen_path = f"img/uploads/{unique}"

        habitacion = nuevaHabitacion(
            nombre=nombre,
            descripcion=descripcion,
            plan=plan,
            numero=numero,
            caracteristicas=caracteristicas,
            precio=precio,
            estado=estado,
            cupo_personas=cupo_personas,
            imagen=imagen_path
        )
        db.session.add(habitacion)
        db.session.commit()

        flash("✅ Habitación creada correctamente", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al crear la habitación: {e}", "danger")

    return redirect(url_for("admin.hospedaje_index"))

#editar habitacion ----------------------------------------------------------

# (El endpoint POST de edición se maneja en 'hospedaje_actualizar')

#eliminar habitacion ----------------------------------------------------------

@admin_bp.route("/hospedaje/eliminar/<int:habitacion_id>", methods=["POST"])
def hospedaje_eliminar(habitacion_id):
    habitacion = nuevaHabitacion.query.get_or_404(habitacion_id)
    try:
        db.session.delete(habitacion)
        db.session.commit()
        flash("🗑️ Habitación eliminada", "warning")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al eliminar: {e}", "danger")

    return redirect(url_for("admin.hospedaje_index"))

# ==========================
# 📌 SECCIÓN RESTAURANTE
# ==========================
_platos_demo = []

@admin_bp.route("/home_dashboard")
def home_dashboard():
    return render_template("dashboard/home_dashboard.html")

@admin_bp.route("/restaurante")
def admin_restaurante():
    return render_template("dashboard/restaurante_admin.html", platos=_platos_demo)

#añadir nuevo plato ---------------------------------------------------------

@admin_bp.route("/restaurante/nuevo", methods=["POST"])
def admin_restaurante_nuevo():
    nombre = request.form.get("nombre")
    categoria = request.form.get("categoria")
    precio = float(request.form.get("precio") or 0)
    nuevo_id = max([p["id"] for p in _platos_demo]) + 1 if _platos_demo else 1
    _platos_demo.append({"id": nuevo_id, "nombre": nombre, "categoria": categoria, "precio": precio})
    return redirect(url_for("admin.admin_restaurante"))

#editar plato ----------------------------------------------------------

@admin_bp.route("/restaurante/editar/<int:plato_id>", methods=["GET", "POST"])
def admin_restaurante_editar(plato_id):
    plato = next((p for p in _platos_demo if p["id"] == plato_id), None)
    if not plato:
        return redirect(url_for("admin.admin_restaurante"))

    if request.method == "POST":
        plato["nombre"] = request.form.get("nombre")
        plato["categoria"] = request.form.get("categoria")
        plato["precio"] = float(request.form.get("precio") or 0)
        return redirect(url_for("admin.admin_restaurante"))

    return render_template("dashboard/restaurante_admin.html", platos=_platos_demo, plato=plato)

#eliminar plato ----------------------------------------------------------

@admin_bp.route("/restaurante/eliminar/<int:plato_id>", methods=["POST"])
def admin_restaurante_eliminar(plato_id):
    global _platos_demo
    _platos_demo = [p for p in _platos_demo if p["id"] != plato_id]

# ==========================
# 📄 NOSOTROS (Contenido dinámico)
# ==========================
@admin_bp.route('/nosotros')
def admin_nosotros():
    posts = Post.query.order_by(Post.creado_en.desc()).all()
    return render_template('dashboard/nosotros_admin.html', posts=posts)

@admin_bp.route('/nosotros/nuevo', methods=['POST'])
def admin_nosotros_nuevo():
    titulo = request.form.get('titulo')
    contenido = request.form.get('contenido')
    categoria = request.form.get('categoria')
    activo = True if request.form.get('activo') == 'on' else False
    imagen = None
    img = request.files.get('imagen')
    if img and img.filename:
        import os, time
        from uuid import uuid4
        from werkzeug.utils import secure_filename
        filename = secure_filename(img.filename)
        unique = f"{int(time.time())}_{uuid4().hex[:8]}_{filename}"
        img_folder = os.path.join(current_app.static_folder, 'img', 'uploads')
        os.makedirs(img_folder, exist_ok=True)
        path = os.path.join(img_folder, unique)
        img.save(path)
        imagen = f"img/uploads/{unique}"
    p = Post(titulo=titulo, contenido=contenido, categoria=categoria, activo=activo, imagen=imagen)
    db.session.add(p)
    db.session.commit()
    flash('Contenido creado', 'success')
    return redirect(url_for('admin.admin_nosotros'))

@admin_bp.route('/nosotros/<int:pid>/editar', methods=['POST'])
def admin_nosotros_editar(pid):
    p = Post.query.get_or_404(pid)
    p.titulo = request.form.get('titulo') or p.titulo
    p.contenido = request.form.get('contenido') or p.contenido
    p.categoria = request.form.get('categoria') or p.categoria
    p.activo = True if request.form.get('activo') == 'on' else False
    img = request.files.get('imagen')
    if img and img.filename:
        import os, time
        from uuid import uuid4
        from werkzeug.utils import secure_filename
        filename = secure_filename(img.filename)
        unique = f"{int(time.time())}_{uuid4().hex[:8]}_{filename}"
        img_folder = os.path.join(current_app.static_folder, 'img', 'uploads')
        os.makedirs(img_folder, exist_ok=True)
        path = os.path.join(img_folder, unique)
        img.save(path)
        p.imagen = f"img/uploads/{unique}"
    db.session.commit()
    flash('Contenido actualizado', 'success')
    return redirect(url_for('admin.admin_nosotros'))

@admin_bp.route('/nosotros/<int:pid>/eliminar', methods=['POST'])
def admin_nosotros_eliminar(pid):
    p = Post.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    flash('Contenido eliminado', 'warning')
    return redirect(url_for('admin.admin_nosotros'))
